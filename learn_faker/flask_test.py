# -*- coding: utf-8 -*-
from flask import Flask, render_template
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')
def index():
    # 加载Excel文件并读取数据
    wb = load_workbook('test.xlsx')
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2):
        data.append([cell.value for cell in row])

    # 渲染HTML模板并返回结果
    return render_template('index.html', data=data)

if __name__ == '__main__':
    app.run(debug=True)
