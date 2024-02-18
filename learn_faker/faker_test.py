# -*- coding: utf-8 -*-
from faker import Faker
from openpyxl import Workbook

# 创建一个workbook对象
wb = Workbook()

# 获取当前活动的工作表
ws = wb.active

# 写入数据到单元格
ws['A1'] = '申请单号'
ws['B1'] = '价税合计'
# ws['C1'] = '单据类型'
# ws['D1'] = '销方名称'
# ws['E1'] = '购方名称'
# ws['F1'] = '发票类型'
# ws['G1'] = '公司代码'

# 使用faker库生成随机数据写入Excel中
fake = Faker("zh_CN")
for i in range(1, 100):
    ws.cell(row=i, column=1).value = fake.pystr()
    ws.cell(row=i, column=2).value = fake.pyfloat()

# 保存Excel
wb.save('test.xlsx')